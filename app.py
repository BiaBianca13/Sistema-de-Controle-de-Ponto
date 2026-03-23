import os
import sqlite3
import openpyxl
import re
import psycopg2.extras
from datetime import datetime, date

from flask import Flask, render_template, request, redirect, flash

app = Flask(__name__)
app.secret_key = "secret_assertividade_total"

# ==========================================
# CONFIGURAÇÃO DE CAMINHOS
# ==========================================
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DB = os.path.join(BASE_DIR, "ponto.db")
DATABASE_URL = os.environ.get("DATABASE_URL")

if DATABASE_URL:
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://")

# ==========================================
# FILTROS E UTILITÁRIOS
# ==========================================
@app.template_filter("tempo")
def tempo(minutos):
    if minutos is None or minutos == 0:
        return "0 min"
    minutos = int(minutos)
    sinal = "-" if minutos < 0 else ""
    minutos = abs(minutos)
    horas = minutos // 60
    resto = minutos % 60
    if horas == 0:
        return f"{sinal}{resto} min"
    if resto == 0:
        return f"{sinal}{horas}h"
    return f"{sinal}{horas}h {resto} min"

def conectar():
    if DATABASE_URL:
        conn = psycopg2.connect(DATABASE_URL)
        conn.cursor_factory = psycopg2.extras.RealDictCursor
        return conn
    else:
        conn = sqlite3.connect(DB)
        conn.row_factory = sqlite3.Row
        return conn

# ==========================================
# LÓGICA DE CÁLCULO
# ==========================================
def extrair_horas(texto):
    if not texto: return []
    linhas = str(texto).split("\n")
    horas = []
    for l in linhas:
        l = l.strip()
        if re.match(r"\d{2}:\d{2}", l):
            horas.append(l)
    return horas

def analisar_batidas(batidas_str, nome, ent_cfg_str, sai_cfg_str, almoco_min, dia_semana, tipo_contrato):
    if not ent_cfg_str: ent_cfg_str = "08:00"
    if not sai_cfg_str: sai_cfg_str = "17:30"
    if not almoco_min: almoco_min = 60

    ent_cfg = datetime.strptime(ent_cfg_str, "%H:%M")
    sai_cfg = datetime.strptime(sai_cfg_str, "%H:%M")
    
    # Horário de Sexta-feira
    if dia_semana == 4 and sai_cfg_str == "17:30":
        sai_cfg = datetime.strptime("17:00", "%H:%M")
        
    # JORNADA ESPERADA
    if dia_semana >= 5: # Fim de semana não tem jornada esperada (Obrigatório = 0)
        esperado = 0
    else:
        if tipo_contrato == "Estagiário":
            esperado = 360 # 6 horas exatas
        else:
            esperado = int((sai_cfg - ent_cfg).total_seconds() / 60) - almoco_min
    
    if not batidas_str:
        falta_real = 1 if dia_semana < 5 else 0
        return 0, esperado, falta_real, "Falta" if falta_real else ""
        
    horarios = sorted([datetime.strptime(h, "%H:%M") for h in batidas_str])
    
    # Filtro anti-duplo clique
    filtrado = [horarios[0]]
    for h in horarios[1:]:
        if (h - filtrado[-1]).total_seconds() / 60 > 10:
            filtrado.append(h)
    horarios = filtrado
    qtd = len(horarios)
    
    if qtd >= 2:
        bruto = (horarios[-1] - horarios[0]).total_seconds() / 60
        intervalo_real = 0
        if qtd >= 4:
            intervalo_real = (horarios[2] - horarios[1]).total_seconds() / 60
        
        desconto = max(intervalo_real, almoco_min)
        trabalhado = int(max(0, bruto - desconto))
        obs = "Sábado/Domingo" if dia_semana >= 5 else "OK"
    else:
        trabalhado = 0
        obs = "Batida incompleta"
        
    return trabalhado, esperado, 0, obs

def ler_excel(caminho):
    wb = openpyxl.load_workbook(caminho)
    ws = wb.active
    dados = []
    periodo = None

    for col in range(1, 10):
        valor = ws.cell(2, col).value
        if valor and "/" in str(valor):
            periodo = valor
            break

    if periodo:
        datas = re.findall(r"\d{2}/\d{2}/\d{4}", str(periodo))
        dia, mes, ano = datas[0].split("/")
        mes, ano = int(mes), int(ano)
    else:
        hoje = date.today()
        mes, ano = hoje.month, hoje.year

    for r in range(1, ws.max_row):
        if ws.cell(r, 1).value == "ID":
            nome = ws.cell(r, 12).value
            linha_dias, linha_horas = r + 1, r + 3
            for col in range(1, 32):
                dia = ws.cell(linha_dias, col).value
                if isinstance(dia, int):
                    batidas = extrair_horas(ws.cell(linha_horas, col).value)
                    dados.append({
                        "nome": nome,
                        "data": date(ano, mes, dia),
                        "batidas": batidas
                    })
    return dados

def criar_banco():
    conn = conectar()
    cur = conn.cursor()
    cur.execute("CREATE TABLE IF NOT EXISTS funcionarios(id INTEGER PRIMARY KEY AUTOINCREMENT, nome TEXT UNIQUE, entrada TEXT, saida TEXT, almoco INTEGER, valor_hora REAL)")
    try:
        cur.execute("ALTER TABLE funcionarios ADD COLUMN tipo_contrato TEXT DEFAULT 'CLT'")
    except:
        pass
    cur.execute("CREATE TABLE IF NOT EXISTS registros(id INTEGER PRIMARY KEY AUTOINCREMENT, funcionario TEXT, data TEXT, batidas TEXT, minutos_trabalhados INTEGER, minutos_esperados INTEGER, falta INTEGER, observacao TEXT)")
    conn.commit()
    conn.close()

criar_banco()

# ==========================================
# ROTAS DO SISTEMA
# ==========================================
@app.route("/")
def dashboard():
    conn = conectar()
    cur = conn.cursor()
    cur.execute("SELECT * FROM registros ORDER BY funcionario, data")
    dados = cur.fetchall()
    conn.close()
    
    resumo = {}
    for d in dados:
        nome = d["funcionario"]
        if nome not in resumo:
            resumo[nome] = {"trabalhado": 0, "esperado": 0, "faltas": 0, "saldo": 0}
        
        resumo[nome]["trabalhado"] += d["minutos_trabalhados"]
        resumo[nome]["esperado"] += d["minutos_esperados"]
        resumo[nome]["faltas"] += d["falta"]
        
        # Saldo do painel ignora o Sábado (5) e Domingo (6)
        dt = datetime.strptime(d["data"], "%Y-%m-%d").date()
        if dt.weekday() < 5:
            resumo[nome]["saldo"] += (d["minutos_trabalhados"] - d["minutos_esperados"])
            
    return render_template("dashboard.html", resumo=resumo)

@app.route("/upload", methods=["GET", "POST"])
def upload():
    if request.method == "POST":
        arquivo = request.files.get("arquivo")
        if not arquivo: return redirect("/upload")
        
        os.makedirs(os.path.join(BASE_DIR, "uploads"), exist_ok=True)
        caminho = os.path.join(BASE_DIR, "uploads", arquivo.filename)
        arquivo.save(caminho)
        registros = ler_excel(caminho)
        
        conn = conectar(); cur = conn.cursor()
        cur.execute("DELETE FROM registros") 
        for r in registros:
            cur.execute("SELECT * FROM funcionarios WHERE UPPER(nome)=UPPER(?)", (r["nome"],))
            f = cur.fetchone()
            ent = f["entrada"] if f else "08:00"
            sai = f["saida"] if f else "17:30"
            alm = f["almoco"] if f else 60
            tipo = dict(f).get("tipo_contrato", "CLT") if f else "CLT"
            
            min_t, min_e, falta, obs = analisar_batidas(r["batidas"], r["nome"], ent, sai, alm, r["data"].weekday(), tipo)
            
            cur.execute("INSERT INTO registros(funcionario, data, batidas, minutos_trabalhados, minutos_esperados, falta, observacao) VALUES (?,?,?,?,?,?,?)",
                (r["nome"], str(r["data"]), "\n".join(r["batidas"]), min_t, min_e, falta, obs))
        conn.commit(); conn.close()
        flash("Planilha processada com sucesso!")
        return redirect("/")
    return render_template("upload.html")

@app.route("/relatorio/<nome>")
def relatorio_individual(nome):
    conn = conectar(); cur = conn.cursor()
    cur.execute("SELECT * FROM registros WHERE funcionario = ? ORDER BY data", (nome,))
    registros = cur.fetchall(); conn.close()
    
    semanas = {}
    total_geral_positivo = 0
    total_geral_negativo = 0
    total_esperado_mes = 0
    total_trabalhado_mes = 0

    for r in registros:
        dt = datetime.strptime(r["data"], "%Y-%m-%d").date()
        
        if dt.weekday() == 6: continue # Remove Domingo

        chave = f"{dt.isocalendar()[0]}-W{dt.isocalendar()[1]}"
        if chave not in semanas:
            semanas[chave] = {"dias": [], "horas_positivas": 0, "horas_negativas": 0}

        total_esperado_mes += r["minutos_esperados"]
        total_trabalhado_mes += r["minutos_trabalhados"]

        if dt.weekday() == 5: # Sábado
            saldo_diario = 0
            obs = "Remunerado por fora" if r["minutos_trabalhados"] > 0 else "Sábado"
        else:
            saldo_diario = r["minutos_trabalhados"] - r["minutos_esperados"]
            obs = r["observacao"]

        semanas[chave]["dias"].append({
            "data": dt.strftime("%d/%m/%Y"),
            "dia_semana": ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb"][dt.weekday()],
            "batidas": r["batidas"].replace("\n", " - "),
            "trabalhado": r["minutos_trabalhados"],
            "saldo_diario": saldo_diario,
            "falta": r["falta"],
            "observacao": obs,
            "is_sabado": True if dt.weekday() == 5 else False
        })
        
        if saldo_diario > 0:
            semanas[chave]["horas_positivas"] += saldo_diario
            total_geral_positivo += saldo_diario
        elif saldo_diario < 0:
            semanas[chave]["horas_negativas"] += abs(saldo_diario)
            total_geral_negativo += abs(saldo_diario)

    for ch, d in semanas.items():
        d["saldo_semana"] = d["horas_positivas"] - d["horas_negativas"]

    total_restante = total_geral_positivo - total_geral_negativo

    return render_template("relatorio.html", nome=nome, semanas=semanas, 
                           total_positivo=total_geral_positivo, 
                           total_negativo=total_geral_negativo, 
                           total_restante=total_restante,
                           total_esperado_mes=total_esperado_mes,
                           total_trabalhado_mes=total_trabalhado_mes)

@app.route("/funcionarios")
def funcionarios():
    conn = conectar(); cur = conn.cursor()
    cur.execute("SELECT * FROM funcionarios"); lista = cur.fetchall(); conn.close()
    return render_template("funcionarios.html", lista=lista)

@app.route("/funcionarios/add", methods=["POST"])
def add_func():
    nome = request.form["nome"]
    ent = request.form["entrada"]
    sai = request.form["saida"]
    alm = request.form["almoco"]
    tipo = request.form.get("tipo_contrato", "CLT")
    
    conn = conectar(); cur = conn.cursor()
    try:
        cur.execute("INSERT INTO funcionarios (nome,entrada,saida,almoco,valor_hora,tipo_contrato) VALUES(?,?,?,?,?,?)", (nome,ent,sai,alm,0,tipo))
    except: pass
    conn.commit(); conn.close()
    flash(f"Funcionário {nome} ({tipo}) adicionado com sucesso!")
    return redirect("/funcionarios")

@app.route("/funcionarios/delete/<int:id>")
def deletar_funcionario(id):
    conn = conectar(); cur = conn.cursor()
    cur.execute("DELETE FROM funcionarios WHERE id=?", (id,))
    conn.commit(); conn.close()
    return redirect("/funcionarios")

# ==========================================
# GERENCIAMENTO DE UPLOADS (ARQUIVOS EXCEL)
# ==========================================
@app.route("/gerenciar-uploads")
def gerenciar_uploads():
    pasta_uploads = os.path.join(BASE_DIR, "uploads")
    os.makedirs(pasta_uploads, exist_ok=True)
    
    arquivos = []
    for nome_arquivo in os.listdir(pasta_uploads):
        caminho_completo = os.path.join(pasta_uploads, nome_arquivo)
        if os.path.isfile(caminho_completo):
            tamanho_kb = os.path.getsize(caminho_completo) / 1024
            data_mod = os.path.getmtime(caminho_completo)
            data_formatada = datetime.fromtimestamp(data_mod).strftime('%d/%m/%Y %H:%M')
            
            arquivos.append({
                "nome": nome_arquivo,
                "tamanho": round(tamanho_kb, 1),
                "data_formatada": data_formatada,
                "timestamp": data_mod
            })
            
    arquivos.sort(key=lambda x: x["timestamp"], reverse=True)
    return render_template("gerenciar_uploads.html", arquivos=arquivos)

@app.route("/deletar-upload/<nome_arquivo>")
def deletar_upload(nome_arquivo):
    caminho_completo = os.path.join(BASE_DIR, "uploads", nome_arquivo)
    if os.path.exists(caminho_completo):
        os.remove(caminho_completo)
        flash(f"Arquivo {nome_arquivo} excluído com sucesso!")
    return redirect("/gerenciar-uploads")

@app.route("/processar-upload/<nome_arquivo>")
def processar_upload(nome_arquivo):
    caminho_completo = os.path.join(BASE_DIR, "uploads", nome_arquivo)
    if os.path.exists(caminho_completo):
        registros = ler_excel(caminho_completo)
        
        conn = conectar(); cur = conn.cursor()
        cur.execute("DELETE FROM registros") 
        for r in registros:
            cur.execute("SELECT * FROM funcionarios WHERE UPPER(nome)=UPPER(?)", (r["nome"],))
            f = cur.fetchone()
            ent = f["entrada"] if f else "08:00"
            sai = f["saida"] if f else "17:30"
            alm = f["almoco"] if f else 60
            tipo = dict(f).get("tipo_contrato", "CLT") if f else "CLT"
            
            min_t, min_e, falta, obs = analisar_batidas(r["batidas"], r["nome"], ent, sai, alm, r["data"].weekday(), tipo)
            
            cur.execute("INSERT INTO registros(funcionario, data, batidas, minutos_trabalhados, minutos_esperados, falta, observacao) VALUES (?,?,?,?,?,?,?)",
                (r["nome"], str(r["data"]), "\n".join(r["batidas"]), min_t, min_e, falta, obs))
        conn.commit(); conn.close()
        
        flash(f"Planilha {nome_arquivo} carregada no sistema com sucesso!")
    return redirect("/")

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)